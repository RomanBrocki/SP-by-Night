import json
import re
import unicodedata
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "tools" / "sp_by_night_source.json"
OUT = ROOT / "tools" / "sp_kindred_residence.json"


def _norm(s: str) -> str:
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _region_to_residence_id(region: str, sect: str) -> tuple[str, str]:
    r = _norm(region)
    if re.search(r"mooca|tatuape|baronato mooca", r):
        return "baronato_mooca_tatuape", "Baronato Mooca/Tatuape"
    if re.search(r"itaquera|leste|guaianases|sapopemba|sao mateus|cidade tiradentes", r):
        return "baronato_leste", "Baronato Leste (Itaquera/Extremo Leste)"
    if re.search(r"capao|grajau|baronato sul|extremo sul|bordas verdes|jardim angela|campo limpo|jardim sao luis", r):
        return "baronato_sul", "Baronato Sul (Capao/Grajau)"
    if re.search(r"centro velho|se republica|\bse\b|republica", r):
        return "indep_centro_velho", "Independentes: Centro Velho"
    if re.search(r"cemiterio|consolacao", r):
        return "indep_cemiterios", "Independentes: Cemiterios/Funerarias (enclave)"
    if re.search(r"barra funda|lapa|rodovia|rodoanel|trajeto|zona cinza|corredor|eixo", r):
        return "indep_corredores", "Independentes: Corredores"

    s = _norm(sect)
    if "anarch" in s or "anarqu" in s:
        return "baronato_mooca_tatuape", "Baronato Mooca/Tatuape"
    if "independ" in s:
        return "indep_centro_velho", "Independentes: Centro Velho"
    return "camarilla_macro", "Camarilla (macro)"


def _load_canon_sheet() -> dict[str, dict]:
    xlsx = next(ROOT.glob("*local*.xlsx"), None)
    if not xlsx:
        raise SystemExit("missing referencia_localizacao.xlsx")
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if "npcs" not in wb.sheetnames:
        raise SystemExit("sheet 'npcs' not found in referencia_localizacao.xlsx")
    ws = wb["npcs"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h): i + 1 for i, h in enumerate(headers)}
    required = {"nome", "regiao", "grupo"}
    if not required.issubset(idx.keys()):
        raise SystemExit("sheet 'npcs' missing required columns: nome, regiao, grupo")

    by_name: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, idx["nome"]).value
        if not nome:
            continue
        regiao = ws.cell(r, idx["regiao"]).value or ""
        grupo = ws.cell(r, idx["grupo"]).value or ""
        cla = ws.cell(r, idx.get("cla_ou_tipo", 0)).value if idx.get("cla_ou_tipo") else ""
        by_name[_norm(str(nome))] = {
            "nome": str(nome),
            "regiao": str(regiao),
            "grupo": str(grupo),
            "cla_ou_tipo": str(cla or ""),
        }
    return by_name


def main() -> int:
    if not SRC.exists():
        raise SystemExit(f"missing {SRC}")
    data = json.loads(SRC.read_text(encoding="utf-8"))
    entities = [e for e in (data.get("entities") or []) if isinstance(e, dict)]
    kindred = [e for e in entities if e.get("kind") == "kindred"]
    canon_by_name = _load_canon_sheet()

    out = {
        "meta": {
            "note": "Residencia canonica dos Cainitas. Fonte unica: referencia_localizacao.xlsx (aba npcs).",
        },
        "residence": {},
    }

    matched = 0
    for e in kindred:
        eid = str(e.get("id") or "").strip()
        if not eid:
            continue
        name = str(e.get("display_name") or e.get("name") or "").strip()
        sect = str(e.get("sect") or "").strip()
        domain_raw = str(e.get("domain") or "").strip()
        canon = canon_by_name.get(_norm(name))
        if canon:
            rid, rlabel = _region_to_residence_id(canon.get("regiao") or "", sect or canon.get("grupo") or "")
            out["residence"][eid] = {
                "name": name,
                "sect": sect,
                "residence_id": rid,
                "residence_label": rlabel,
                "source": "referencia_localizacao.xlsx:npcs",
                "domain_raw": domain_raw,
                "canon_region": canon.get("regiao") or "",
                "canon_group": canon.get("grupo") or "",
            }
            matched += 1
        else:
            rid, rlabel = _region_to_residence_id("", sect)
            out["residence"][eid] = {
                "name": name,
                "sect": sect,
                "residence_id": rid,
                "residence_label": rlabel,
                "source": "fallback_by_sect",
                "domain_raw": domain_raw,
                "canon_region": "",
                "canon_group": "",
            }

    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT} (matched {matched}/{len(kindred)})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
