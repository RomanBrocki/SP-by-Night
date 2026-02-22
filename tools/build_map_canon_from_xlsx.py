from __future__ import annotations

import json
import math
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUT_06 = ROOT / "06_MAPA_SP" / "data" / "canon_map_data.json"
OUT_DOCS = ROOT / "docs" / "map" / "data" / "canon_map_data.json"
DISTRICTS_GEOJSON = ROOT / "06_MAPA_SP" / "data" / "distritos-sp.geojson"
BOOK_DATA_JSON = ROOT / "07_LIVRO_BY_NIGHT" / "book_data.json"
BARONATOS_OVERLAY_JSON = ROOT / "06_MAPA_SP" / "data" / "baronatos_overlay_editor.json"
SOURCE_JSON = ROOT / "tools" / "sp_by_night_source.json"


ANARCH_DISTRICTS = [
    "BELEM",
    "MOOCA",
    "TATUAPE",
    "CIDADE LIDER",
    "ITAQUERA",
    "JOSE BONIFACIO",
    "PARQUE DO CARMO",
    "VILA JACUI",
    "CAMPO LIMPO",
    "CAPAO REDONDO",
    "CIDADE DUTRA",
    "GRAJAU",
    "JARDIM ANGELA",
    "JARDIM SAO LUIS",
    "SOCORRO",
]

INDEP_DISTRICTS = ["SE", "REPUBLICA", "CONSOLACAO"]

LUPINO_DISTRICTS = [
    "ANHANGUERA",
    "PERUS",
    "JARAGUA",
    "TREMEMBE",
    "JACANA",
    "PARELHEIROS",
    "MARSILAC",
]

ANARCH_MOOCA_TATUAPE = ["MOOCA", "BELEM", "TATUAPE"]
ANARCH_LESTE = ["CIDADE LIDER", "ITAQUERA", "JOSE BONIFACIO", "PARQUE DO CARMO", "VILA JACUI"]
ANARCH_SUL = ["CAPAO REDONDO", "GRAJAU", "JARDIM ANGELA", "JARDIM SAO LUIS", "CAMPO LIMPO", "SOCORRO", "CIDADE DUTRA"]

CAMARILLA_REGION_DISTRICTS: list[tuple[str, list[str]]] = [
    (r"morumbi|campo belo", ["MORUMBI", "CAMPO BELO", "VILA ANDRADE"]),
    (r"itaim|berrini", ["ITAIM BIBI", "SANTO AMARO", "CAMPO BELO"]),
    (r"paulista|jardins|jardim paulista", ["BELA VISTA", "JARDIM PAULISTA", "CONSOLACAO"]),
    (r"pinheiros|vila madalena|baixa augusta", ["PINHEIROS", "ALTO DE PINHEIROS", "CONSOLACAO"]),
    (r"moema|vila mariana|paraiso", ["MOEMA", "VILA MARIANA", "SAUDE"]),
    (r"santana|tucuruvi", ["SANTANA", "TUCURUVI", "MANDAQUI"]),
    (r"ipiranga|cursino", ["IPIRANGA", "CURSINO", "SACOMA"]),
    (r"bela vista|bixiga", ["BELA VISTA", "CONSOLACAO", "LIBERDADE"]),
    (r"butanta|usp|alto de pinheiros", ["BUTANTA", "RIO PEQUENO", "ALTO DE PINHEIROS"]),
    (r"bras|pari", ["BRAS", "PARI", "MOOCA"]),
    (r"perdizes|barra funda", ["PERDIZES", "BARRA FUNDA", "LAPA"]),
    (r"consolacao|higienopolis", ["CONSOLACAO", "SANTA CECILIA", "PERDIZES"]),
    (r"liberdade", ["LIBERDADE", "SE", "BELA VISTA"]),
    (r"centro", ["SE", "REPUBLICA", "BELA VISTA", "CONSOLACAO", "LIBERDADE"]),
]

INDEPENDENT_REGION_DISTRICTS: list[tuple[str, list[str]]] = [
    (r"sul|itapecerica|bordas verdes", ["PARELHEIROS", "MARSILAC", "PEDREIRA", "CIDADE ADEMAR", "JABAQUARA", "SANTO AMARO"]),
    (r"pinheiros|baixa augusta", ["PINHEIROS", "ALTO DE PINHEIROS", "CONSOLACAO", "BELA VISTA"]),
    (r"oeste|perdizes|barra funda|lapa", ["PERDIZES", "BARRA FUNDA", "LAPA", "VILA LEOPOLDINA", "JAGUARA", "PINHEIROS"]),
    (r"eixos urbanos|estacoes|galerias", ["BARRA FUNDA", "LAPA", "BRAS", "PARI", "PINHEIROS", "TATUAPE", "MOOCA"]),
    (r"centro|clinicas clandestinas", ["BARRA FUNDA", "SANTA CECILIA", "BELA VISTA", "LIBERDADE", "BRAS", "PARI"]),
    (r"centro/barra funda", ["BARRA FUNDA", "SANTA CECILIA", "PERDIZES", "LAPA", "LIBERDADE"]),
    (r"rodoanel|rodovias|aeroporto|trajetos|eixos de carga|entrega", ["JABAQUARA", "SANTO AMARO", "LAPA", "BARRA FUNDA", "ANHANGUERA", "PERUS", "PARELHEIROS", "MARSILAC"]),
]

# Regras canonicas explicitadas pelo curador para residencia/territorio.
FORCED_DISTRICT_BY_NAME: dict[str, list[str]] = {
    # Tremere: Capela no Butanta; Dario com territorio no Butanta.
    "dario kron": ["BUTANTA"],
    # Garantir ao menos 3 Tremere residentes na capela.
    "caio menezes": ["BUTANTA"],
    "igor menezes": ["BUTANTA"],
    "maira koehler": ["BUTANTA"],
    # Hecata: lider + mais um na Se; demais em nucleos de cemiterio.
    "donato lazzari": ["SE"],
    "soraia nunes": ["SE"],
    "celia moura": ["CONSOLACAO"],
    "iago siqueira": ["REPUBLICA"],
    # Independente Tzimisce no Sul, fora do nucleo Hecata.
    "vlado de itapecerica": ["PARELHEIROS", "PEDREIRA", "MARSILAC"],
}

# Coordenadas canônicas fixas para pontos-chave em mapa.
# Mantém legibilidade visual e evita deriva para áreas vazias dentro do distrito.
FORCED_COORD_BY_NAME: dict[str, tuple[float, float]] = {
    # Hecata líder + núcleo na Sé.
    "donato lazzari": (-23.5505, -46.6345),
    "soraia nunes": (-23.5511, -46.6337),
    # Capela Tremere no Butantã (núcleo principal).
    "dario kron": (-23.5652, -46.7140),
    "igor menezes": (-23.5663, -46.7127),
    "caio menezes": (-23.5642, -46.7163),
    "maira koehler": (-23.5671, -46.7156),
    # Independente Tzimisce no sul urbano da Pedreira.
    "vlado de itapecerica": (-23.7060, -46.6470),
}

def norm(s: str) -> str:
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def slug(s: str) -> str:
    s = norm(s)
    return s.replace(" ", "_")


def stem_from_name(name: str) -> str:
    s = unicodedata.normalize("NFKD", str(name or "")).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_")
    return re.sub(r"_+", "_", s)


def canon_name_key(name: str) -> str:
    return norm(name)


def fix_mojibake(s: str) -> str:
    txt = str(s or "")
    if not txt:
        return ""
    # Some spreadsheet cells were typed with look-alike glyphs.
    txt = (
        txt.replace("\u0103", "\u00e3")
        .replace("\u0102", "\u00c3")
        .replace("\u021b", "t")
        .replace("\u021a", "T")
        .replace("\u0219", "s")
        .replace("\u0218", "S")
        .replace("\u0163", "t")
        .replace("\u015f", "s")
    )
    # Try common latin1->utf8 recovery when text shows mojibake markers.
    if any(ch in txt for ch in ("\u00c3", "\u00e2", "\u00c2")):
        try:
            fixed = txt.encode("latin1", errors="ignore").decode("utf-8", errors="ignore")
            if fixed.strip():
                return fixed
        except Exception:
            pass
    return txt


def load_role_index_from_book_data() -> dict[str, dict[str, str]]:
    if not BOOK_DATA_JSON.exists():
        return {}
    try:
        data = json.loads(BOOK_DATA_JSON.read_text(encoding="utf-8"))
    except Exception:
        return {}
    entities = data.get("entities") or []
    out: dict[str, dict[str, str]] = {}
    for e in entities:
        display_name = str(e.get("display_name") or "").strip()
        if not display_name:
            continue
        file_stem = str(e.get("file_stem") or "").strip()
        out[canon_name_key(display_name)] = {
            "role": fix_mojibake(str(e.get("role") or "").strip()),
            "domain_text": fix_mojibake(str(e.get("domain") or "").strip()),
            "file_stem": file_stem,
        }
    return out


def load_ghoul_master_map_from_source() -> dict[str, str]:
    if not SOURCE_JSON.exists():
        return {}
    try:
        data = json.loads(SOURCE_JSON.read_text(encoding="utf-8"))
    except Exception:
        return {}
    entities = data.get("entities") or []
    by_id: dict[str, dict[str, Any]] = {}
    for e in entities:
        eid = str(e.get("id") or "").strip()
        if eid:
            by_id[eid] = e

    out: dict[str, str] = {}
    for e in entities:
        kind = norm(str(e.get("kind") or ""))
        if kind != "ghoul":
            continue
        g_name = str(e.get("display_name") or "").strip()
        if not g_name:
            continue
        links = e.get("links") or []
        preferred: str | None = None
        fallback: str | None = None
        for lk in links:
            lk_type = norm(str(lk.get("type") or ""))
            to_id = str(lk.get("to") or "").strip()
            t = by_id.get(to_id) or {}
            t_kind = norm(str(t.get("kind") or ""))
            t_name = str(t.get("display_name") or "").strip()
            if not t_name or t_kind != "kindred":
                continue
            if lk_type == "servant":
                preferred = t_name
                break
            if lk_type == "canon anchor" and fallback is None:
                fallback = t_name
        m_name = preferred or fallback
        if m_name:
            out[canon_name_key(g_name)] = m_name
    return out


def load_baronatos_overlay() -> None:
    global ANARCH_DISTRICTS, ANARCH_MOOCA_TATUAPE, ANARCH_LESTE, ANARCH_SUL
    if not BARONATOS_OVERLAY_JSON.exists():
        return
    try:
        raw = json.loads(BARONATOS_OVERLAY_JSON.read_text(encoding="utf-8"))
    except Exception:
        return
    b = raw.get("baronatos") or {}
    ferr = [str(x).replace("_", " ").upper() for x in (b.get("ferrugem") or [])]
    leste = [str(x).replace("_", " ").upper() for x in (b.get("leste_aco") or [])]
    sul = [str(x).replace("_", " ").upper() for x in (b.get("sul") or [])]
    if ferr:
        ANARCH_MOOCA_TATUAPE = ferr
    if leste:
        ANARCH_LESTE = leste
    if sul:
        ANARCH_SUL = sul
    union = []
    for arr in [ANARCH_MOOCA_TATUAPE, ANARCH_LESTE, ANARCH_SUL]:
        for d in arr:
            if d not in union:
                union.append(d)
    if union:
        ANARCH_DISTRICTS = union


def point_in_ring(x: float, y: float, ring: list[list[float]]) -> bool:
    inside = False
    n = len(ring)
    for i in range(n):
        x1, y1 = ring[i][0], ring[i][1]
        x2, y2 = ring[(i + 1) % n][0], ring[(i + 1) % n][1]
        if ((y1 > y) != (y2 > y)) and (x < (x2 - x1) * (y - y1) / ((y2 - y1) + 1e-12) + x1):
            inside = not inside
    return inside


def point_in_polygon(x: float, y: float, polycoords: list[list[list[float]]]) -> bool:
    if not polycoords:
        return False
    if not point_in_ring(x, y, polycoords[0]):
        return False
    for hole in polycoords[1:]:
        if point_in_ring(x, y, hole):
            return False
    return True


def point_in_geometry(x: float, y: float, geom: dict[str, Any]) -> bool:
    gtype = geom.get("type")
    coords = geom.get("coordinates")
    if gtype == "Polygon":
        return point_in_polygon(x, y, coords)
    if gtype == "MultiPolygon":
        return any(point_in_polygon(x, y, poly) for poly in coords)
    return False


def geometry_centroid(geom: dict[str, Any]) -> tuple[float, float] | None:
    gtype = geom.get("type")
    coords = geom.get("coordinates") or []
    pts: list[tuple[float, float]] = []
    if gtype == "Polygon":
        if coords and coords[0]:
            pts = [(float(x), float(y)) for x, y in coords[0]]
    elif gtype == "MultiPolygon":
        for poly in coords:
            if poly and poly[0]:
                pts.extend((float(x), float(y)) for x, y in poly[0])
    if not pts:
        return None
    cx = sum(p[0] for p in pts) / len(pts)
    cy = sum(p[1] for p in pts) / len(pts)
    return (cy, cx)


def load_district_geometries() -> dict[str, dict[str, Any]]:
    gj = json.loads(DISTRICTS_GEOJSON.read_text(encoding="utf-8"))
    out: dict[str, dict[str, Any]] = {}
    for ft in gj.get("features", []):
        props = ft.get("properties") or {}
        name = str(props.get("ds_nome") or props.get("name") or "").strip()
        if not name:
            continue
        geom = ft.get("geometry") or {}
        centroid = geometry_centroid(geom)
        out[dk(name)] = {
            "name": name,
            "geometry": geom,
            "centroid": centroid,
        }
    return out


def dk(s: str) -> str:
    return norm(s).upper().replace(" ", "_")


def district_key_set(names: list[str]) -> set[str]:
    return {dk(n) for n in names}


def filter_existing_districts(keys: set[str], district_index: dict[str, dict[str, Any]]) -> set[str]:
    return {k for k in keys if k in district_index}


def districts_for_pattern_map(
    region: str, pattern_map: list[tuple[str, list[str]]], district_index: dict[str, dict[str, Any]]
) -> set[str]:
    r = norm(region)
    found: set[str] = set()
    for patt, names in pattern_map:
        if re.search(patt, r):
            found |= filter_existing_districts(district_key_set(names), district_index)
    return found


def allowed_districts_for_npc(
    region: str,
    faction: str,
    cluster: str,
    clan: str,
    district_index: dict[str, dict[str, Any]],
) -> set[str]:
    r = norm(region)
    c = norm(cluster)
    fac = norm(faction)
    cl = norm(clan)

    if fac == "anarquistas":
        if re.search(r"mooca|belem|tatuape", r):
            return filter_existing_districts(district_key_set(ANARCH_MOOCA_TATUAPE), district_index)
        if re.search(r"itaquera|leste|cidade lider|jose bonifacio|parque do carmo|parque guarani|vila chavante|vila jacui", r):
            return filter_existing_districts(district_key_set(ANARCH_LESTE), district_index)
        if re.search(r"capao|grajau|jardim angela|jardim sao luis|campo limpo|sul|extremo sul", r):
            return filter_existing_districts(district_key_set(ANARCH_SUL), district_index)
        if "interbaronatos" in c:
            return filter_existing_districts(district_key_set(ANARCH_DISTRICTS), district_index)
        return filter_existing_districts(district_key_set(ANARCH_DISTRICTS), district_index)

    if fac == "independentes":
        # Hecata (e servos de necropolis) ficam no nucleo Centro Velho / Consolacao.
        if cl == "hecata" or "hecata" in c or re.search(r"cemiterio|centro velho|consolacao", r):
            return filter_existing_districts(district_key_set(INDEP_DISTRICTS), district_index)

        # Independentes nao-Hecata nao ocupam o nucleo Hecata (SE/REPUBLICA/CONSOLACAO).
        disallowed = filter_existing_districts(
            district_key_set(ANARCH_DISTRICTS + INDEP_DISTRICTS), district_index
        )
        ds = districts_for_pattern_map(region, INDEPENDENT_REGION_DISTRICTS, district_index)
        if ds:
            ds = {k for k in ds if k not in disallowed}
            if ds:
                return ds

        ds2 = districts_for_pattern_map(region, CAMARILLA_REGION_DISTRICTS, district_index)
        if ds2:
            ds2 = {k for k in ds2 if k not in disallowed}
            if ds2:
                return ds2

        # Fallback: cidade fora dos baronatos e fora do nucleo Hecata.
        return {k for k in district_index.keys() if k not in disallowed}

    if fac == "camarilla":
        if "centro expandido" in r:
            disallowed = filter_existing_districts(
                district_key_set(ANARCH_DISTRICTS + INDEP_DISTRICTS), district_index
            )
            return {k for k in district_index.keys() if k not in disallowed}
        ds = districts_for_pattern_map(region, CAMARILLA_REGION_DISTRICTS, district_index)
        disallowed = filter_existing_districts(
            district_key_set(ANARCH_DISTRICTS + INDEP_DISTRICTS), district_index
        )
        if ds:
            ds = {k for k in ds if k not in disallowed}
            if ds:
                return ds
        # Camarilla domina o resto da cidade fora polos de Anarch/Indep.
        return {k for k in district_index.keys() if k not in disallowed}

    # SI/mortais/outros: sem restricao territorial rigida.
    return set()


def forced_districts_for_npc(
    name: str,
    faction: str,
    clan: str,
    district_index: dict[str, dict[str, Any]],
) -> set[str]:
    key = norm(name)
    forced = FORCED_DISTRICT_BY_NAME.get(key)
    if forced:
        return filter_existing_districts(set(forced), district_index)
    # fallback sem regra forçada
    return set()


def max_spread_radius_km(entry: dict[str, Any]) -> float:
    fac = norm(entry.get("faction", ""))
    typ = norm(entry.get("npc_type", ""))
    clan = norm(entry.get("clan", ""))
    allowed_count = len(entry.get("_allowed_districts") or [])
    if fac == "anarquistas":
        return min(9.0, 5.6 + 0.35 * max(0, allowed_count - 1))
    if fac == "independentes":
        return min(7.5, 4.8 + 0.30 * max(0, allowed_count - 1))
    if fac == "camarilla":
        # Camarilla: espalhar mais dentro do territorio canonico.
        # Excecao de permissao: Hecata/Tremere podem ficar mais proximos.
        if typ == "vampiro":
            base = 4.2 if clan in ("hecata", "tremere") else 6.6
            return min(12.0, base + 0.55 * max(0, allowed_count - 1))
        return min(8.0, 3.2 + 0.35 * max(0, allowed_count - 1))
    return 3.2


def spread_anchor_penalty_coeff(entry: dict[str, Any]) -> float:
    fac = norm(entry.get("faction", ""))
    clan = norm(entry.get("clan", ""))
    typ = norm(entry.get("npc_type", ""))
    if fac == "camarilla" and typ == "vampiro":
        # Mantem permissao de proximidade para Tremere/Hecata.
        if clan in ("hecata", "tremere"):
            return 0.18
        # Restante da Camarilla: dispersao maxima no territorio.
        return 0.04
    if fac == "anarquistas":
        return 0.10
    if fac == "independentes":
        return 0.12
    return 0.16


def preferred_min_separation_km(entry: dict[str, Any]) -> float:
    fac = norm(entry.get("faction", ""))
    clan = norm(entry.get("clan", ""))
    typ = norm(entry.get("npc_type", ""))
    if fac == "camarilla" and typ == "vampiro":
        if clan in ("hecata", "tremere"):
            return 0.25
        return 0.90
    if fac == "anarquistas" and typ == "vampiro":
        return 0.55
    if fac == "independentes" and typ == "vampiro":
        return 0.45
    return 0.30


def _sheet_rows(ws) -> list[dict[str, Any]]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h): i + 1 for i, h in enumerate(headers)}
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(r, idx[str(h)]).value for h in headers}
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return rows


def infer_faction(group: str) -> str:
    g = norm(group)
    if "anarqu" in g:
        return "Anarquistas"
    if "independ" in g:
        return "Independentes"
    if "camarilla" in g:
        return "Camarilla"
    if "segunda inquisicao" in g:
        return "Segunda Inquisicao"
    if "mortal" in g:
        return "Mortal"
    return "Outro"


def infer_npc_type(group: str, clan_or_type: str) -> str:
    c = norm(clan_or_type)
    if c == "mortal":
        f = infer_faction(group)
        if f in ("Mortal", "Segunda Inquisicao"):
            return "Mortal"
        return "Ghoul"
    return "Vampiro"


def anchor_for_region(region: str, faction: str, cluster: str) -> tuple[float, float, str]:
    r = norm(region)
    c = norm(cluster)

    # Forced baronato placement for Anarch canon.
    if faction == "Anarquistas":
        if re.search(r"leste|itaquera|cidade tiradentes|guaianases|sao mateus|sapopemba", r):
            return -23.5405, -46.4701, "baronato_leste"
        if re.search(r"sul|capao|grajau|jardim angela|jardim sao luis|campo limpo|extremo sul", r):
            return -23.6496, -46.7524, "baronato_sul"
        if "interbaronatos" in c:
            # Keep interbaronatos inside declared baronatos: default Mooca/Tatuape.
            return -23.5560, -46.5935, "baronato_mooca_tatuape"
        return -23.5560, -46.5935, "baronato_mooca_tatuape"

    if faction == "Independentes":
        if "hecata" in c or re.search(r"cemiterio|consolacao", r):
            return -23.5510, -46.6570, "indep_cemiterios"
        if re.search(r"se|centro velho|republica", r):
            return -23.5480, -46.6350, "indep_centro_velho"
        if re.search(r"sul|itapecerica|bordas verdes", r):
            return -23.6640, -46.7420, "indep_sul"
        if re.search(r"oeste|perdizes|lapa|barra funda", r):
            return -23.5380, -46.6840, "indep_oeste"
        if re.search(r"pinheiros|baixa augusta", r):
            return -23.5625, -46.6850, "indep_pinheiros"
        if re.search(r"aeroporto|rodovia|rodoanel|trajetos|carga|entrega", r):
            return -23.6110, -46.6740, "indep_corredores"
        return -23.5625, -46.6850, "indep_corredores"


    # Camarilla + mortais + SI
    if re.search(r"morumbi|campo belo", r):
        return -23.5974, -46.7068, "cam_morumbi_campo_belo"
    if re.search(r"itaim|berrini", r):
        return -23.5844, -46.6784, "cam_itaim_berrini"
    if re.search(r"paulista|jardins|jardim paulista", r):
        return -23.5561, -46.6622, "cam_paulista_jardins"
    if re.search(r"pinheiros|vila madalena|baixa augusta", r):
        return -23.5673, -46.7019, "cam_pinheiros_vila_madalena"
    if re.search(r"moema|vila mariana|paraiso", r):
        return -23.5930, -46.6625, "cam_moema_vila_mariana"
    if re.search(r"santana|tucuruvi", r):
        return -23.5025, -46.6249, "cam_santana_tucuruvi"
    if re.search(r"ipiranga|cursino", r):
        return -23.5893, -46.6062, "cam_ipiranga_cursino"
    if re.search(r"bela vista|bixiga", r):
        return -23.5577, -46.6468, "cam_bela_vista_bixiga"
    if re.search(r"butanta|usp|alto de pinheiros", r):
        return -23.5652, -46.7126, "cam_butanta_usp"
    if re.search(r"bras|pari", r):
        return -23.5453, -46.6164, "cam_bras_pari"
    if re.search(r"perdizes|barra funda", r):
        return -23.5380, -46.6807, "cam_perdizes_barra_funda"
    if re.search(r"consolacao|higienopolis", r):
        return -23.5490, -46.6523, "cam_consolacao_higienopolis"
    if re.search(r"liberdade", r):
        return -23.5550, -46.6355, "cam_liberdade"
    if re.search(r"centro", r):
        return -23.5485, -46.6380, "cam_centro_expandido"
    if re.search(r"oeste", r):
        return -23.5500, -46.7000, "cam_oeste"
    return -23.5561, -46.6622, "cam_centro_expandido"


def split_region_variants(region: str) -> list[str]:
    raw = str(region or "").strip()
    if not raw:
        return []
    parts = [p.strip() for p in re.split(r"[\/|]", raw) if str(p or "").strip()]
    out: list[str] = []
    for p in parts:
        if p not in out:
            out.append(p)
    return out


def choose_anchor_for_npc(
    region: str,
    faction: str,
    cluster: str,
    tier: str,
    npc_type: str,
    anchor_load: dict[str, int],
    allowed_districts: set[str],
    district_index: dict[str, dict[str, Any]],
    name: str,
) -> tuple[float, float, str]:
    base = anchor_for_region(region, faction, cluster)
    base_lat, base_lon, base_bucket = base

    # Sem distritos permitidos, mantem ancora textual por regiao.
    if not allowed_districts:
        anchor_load[base_bucket] = anchor_load.get(base_bucket, 0) + 1
        return base

    # Escolhe ancora por distrito permitido (canon), minimizando concentracao.
    candidates: list[tuple[str, float, float]] = []
    for dkey in sorted(allowed_districts):
        d = district_index.get(dkey) or {}
        cent = d.get("centroid")
        if not cent:
            continue
        lat, lon = float(cent[0]), float(cent[1])
        candidates.append((dkey, lat, lon))

    if not candidates:
        anchor_load[base_bucket] = anchor_load.get(base_bucket, 0) + 1
        return base

    name_hash = abs(hash(norm(name))) % 997

    def score(cand: tuple[str, float, float]) -> tuple[float, float, int, str]:
        dkey, lat, lon = cand
        key = f"district:{dkey}"
        load = float(anchor_load.get(key, 0))
        dist = dist_km(lat, lon, base_lat, base_lon)
        # tie-break estavel por nome para evitar sempre o mesmo distrito
        tie = (abs(hash(f"{name_hash}:{dkey}")) % 1000)
        return (load, dist, tie, dkey)

    chosen_key, chosen_lat, chosen_lon = sorted(candidates, key=score)[0]
    load_key = f"district:{chosen_key}"
    anchor_load[load_key] = anchor_load.get(load_key, 0) + 1
    return chosen_lat, chosen_lon, load_key


def place_servants_near_masters(
    npcs: list[dict[str, Any]],
    ghoul_master_map: dict[str, str],
    district_index: dict[str, dict[str, Any]],
) -> None:
    if not ghoul_master_map:
        return
    by_name: dict[str, dict[str, Any]] = {}
    for n in npcs:
        by_name[canon_name_key(str(n.get("name") or ""))] = n

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for n in npcs:
        if norm(str(n.get("npc_type") or "")) != "ghoul":
            continue
        gk = canon_name_key(str(n.get("name") or ""))
        m_name = ghoul_master_map.get(gk)
        if not m_name:
            continue
        mk = canon_name_key(m_name)
        master = by_name.get(mk)
        if not master:
            continue
        if norm(str(master.get("npc_type") or "")) != "vampiro":
            continue
        n["_master_key"] = mk
        groups[mk].append(n)

    for mk, servants in groups.items():
        master = by_name.get(mk)
        if not master:
            continue
        m_lat = float(master.get("lat") or master.get("lat_anchor") or 0.0)
        m_lon = float(master.get("lon") or master.get("lon_anchor") or 0.0)
        m_dk = find_district_key_for_point(m_lat, m_lon, district_index) or ""
        servants.sort(key=lambda e: norm(str(e.get("name") or "")))
        n = len(servants)
        base_angle = (abs(hash(mk)) % 360) * math.pi / 180.0
        for i, e in enumerate(servants):
            allowed = set(e.get("_allowed_districts") or [])
            if m_dk:
                allowed.add(m_dk)
            chosen: tuple[float, float] | None = None
            # Gera candidatos em aneis proximos ao mestre.
            for r_km in [0.12, 0.22, 0.35, 0.50, 0.70]:
                for step in range(max(8, n * 2)):
                    ang = base_angle + (2 * math.pi * ((i + step) % max(8, n * 2)) / max(8, n * 2))
                    dlat = (r_km / 111.0) * math.sin(ang)
                    cos_lat = max(0.25, math.cos(math.radians(m_lat)))
                    dlon = (r_km / (111.0 * cos_lat)) * math.cos(ang)
                    lat = round(m_lat + dlat, 6)
                    lon = round(m_lon + dlon, 6)
                    if allowed and not is_inside_any_district(lat, lon, allowed, district_index):
                        continue
                    chosen = (lat, lon)
                    break
                if chosen is not None:
                    break
            if chosen is None:
                chosen = (round(m_lat, 6), round(m_lon, 6))
            e["lat"] = chosen[0]
            e["lon"] = chosen[1]


def spread_points(entries: list[dict[str, Any]], radius_km: float, seed_shift: float) -> None:
    if not entries:
        return
    n = len(entries)
    golden = math.pi * (3 - math.sqrt(5))
    entries.sort(key=lambda x: norm(x["name"]))
    for i, e in enumerate(entries):
        lat0 = e["lat_anchor"]
        lon0 = e["lon_anchor"]
        theta = i * golden + seed_shift
        r_km = radius_km * math.sqrt((i + 0.7) / max(1.0, n))
        dlat = (r_km / 111.0) * math.sin(theta)
        cos_lat = max(0.25, math.cos(math.radians(lat0)))
        dlon = (r_km / (111.0 * cos_lat)) * math.cos(theta)
        e["lat"] = round(lat0 + dlat, 6)
        e["lon"] = round(lon0 + dlon, 6)


def dist_km(a_lat: float, a_lon: float, b_lat: float, b_lon: float) -> float:
    r = 6371.0
    p1 = math.radians(a_lat)
    p2 = math.radians(b_lat)
    dp = math.radians(b_lat - a_lat)
    dl = math.radians(b_lon - a_lon)
    h = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(max(0.0, min(1.0, h))))


def candidate_points_for_entry(
    entry: dict[str, Any],
    geoms: dict[str, dict[str, Any]],
    max_radius_km: float,
) -> list[tuple[float, float]]:
    lat0 = float(entry["lat_anchor"])
    lon0 = float(entry["lon_anchor"])
    allowed = set(entry.get("_allowed_districts") or [])
    out: list[tuple[float, float]] = []

    rings = 18
    per_ring = 32
    for r_i in range(rings + 1):
        frac = math.sqrt(r_i / max(1, rings))
        r_km = max_radius_km * frac
        for a_i in range(per_ring):
            ang = (2 * math.pi * a_i / per_ring) + (0.37 * r_i)
            dlat = (r_km / 111.0) * math.sin(ang)
            cos_lat = max(0.25, math.cos(math.radians(lat0)))
            dlon = (r_km / (111.0 * cos_lat)) * math.cos(ang)
            lat = lat0 + dlat
            lon = lon0 + dlon
            if allowed:
                if not is_inside_any_district(lat, lon, allowed, geoms):
                    continue
            out.append((round(lat, 6), round(lon, 6)))

    if not out:
        out.append((round(lat0, 6), round(lon0, 6)))
    return out


def disperse_bucket(entries: list[dict[str, Any]], geoms: dict[str, dict[str, Any]]) -> None:
    if not entries:
        return
    entries.sort(key=lambda e: norm(e.get("name", "")))
    placed: list[tuple[float, float]] = []

    for entry in entries:
        rad = max_spread_radius_km(entry)
        cands = candidate_points_for_entry(entry, geoms, max_radius_km=rad)
        best = cands[0]
        best_score = -1e18
        anchor_lat = float(entry["lat_anchor"])
        anchor_lon = float(entry["lon_anchor"])

        for lat, lon in cands:
            if placed:
                min_sep = min(dist_km(lat, lon, p_lat, p_lon) for p_lat, p_lon in placed)
            else:
                min_sep = rad
            anchor_penalty = spread_anchor_penalty_coeff(entry) * dist_km(
                lat, lon, anchor_lat, anchor_lon
            )
            req_sep = preferred_min_separation_km(entry)
            sep_penalty = max(0.0, req_sep - min_sep) * 4.0
            score = min_sep - anchor_penalty - sep_penalty
            if score > best_score:
                best_score = score
                best = (lat, lon)

        entry["lat"] = best[0]
        entry["lon"] = best[1]
        placed.append(best)


def find_district_key_for_point(
    lat: float,
    lon: float,
    district_index: dict[str, dict[str, Any]],
) -> str | None:
    x, y = lon, lat
    for k, d in district_index.items():
        geom = d.get("geometry") or {}
        if geom and point_in_geometry(x, y, geom):
            return k
    return None


def convert_seed(v: Any) -> float | None:
    if v in (None, ""):
        return None
    try:
        x = float(v)
    except Exception:
        return None
    # seeds in spreadsheet use 1e7 integer encoding
    if abs(x) > 1000:
        return x / 1e7
    return x


def pick_readable_xlsx() -> Path | None:
    cands = sorted(ROOT.glob("*local*.xlsx"), key=lambda p: (len(p.name), p.name.lower()))
    primary = [p for p in cands if "copia" not in norm(p.name)]
    for p in primary:
        try:
            with p.open("rb"):
                pass
            return p
        except Exception:
            continue
    return None


def is_inside_any_district(lat: float, lon: float, district_keys: set[str], geoms: dict[str, dict[str, Any]]) -> bool:
    x, y = lon, lat
    for d in district_keys:
        geom = (geoms.get(d) or {}).get("geometry")
        if geom and point_in_geometry(x, y, geom):
            return True
    return False


def snap_into_districts(
    entry: dict[str, Any],
    district_keys: set[str],
    geoms: dict[str, dict[str, Any]],
    fallback_lat: float,
    fallback_lon: float,
) -> None:
    if is_inside_any_district(entry["lat"], entry["lon"], district_keys, geoms):
        return

    # Spiral search around fallback anchor until it lands inside allowed polygons.
    lat0, lon0 = fallback_lat, fallback_lon
    for ring in [0.0000, 0.0030, 0.0060, 0.0100, 0.0150, 0.0200, 0.0300, 0.0400, 0.0500, 0.0600, 0.0800]:
        for i in range(36):
            ang = (2 * math.pi * i) / 36.0
            lat = lat0 + ring * math.sin(ang)
            cos_lat = max(0.25, math.cos(math.radians(lat0)))
            lon = lon0 + (ring / cos_lat) * math.cos(ang)
            if is_inside_any_district(lat, lon, district_keys, geoms):
                entry["lat"] = round(lat, 6)
                entry["lon"] = round(lon, 6)
                return

    # Last fallback: keep anchor.
    entry["lat"] = round(lat0, 6)
    entry["lon"] = round(lon0, 6)


def apply_forced_coords(npcs: list[dict[str, Any]], district_index: dict[str, dict[str, Any]]) -> None:
    for n in npcs:
        key = norm(str(n.get("name") or ""))
        forced = FORCED_COORD_BY_NAME.get(key)
        if not forced:
            continue
        lat, lon = float(forced[0]), float(forced[1])
        allowed = set(n.get("_allowed_districts") or [])
        temp = {"lat": lat, "lon": lon}
        if allowed and not is_inside_any_district(lat, lon, allowed, district_index):
            snap_into_districts(
                temp,
                allowed,
                district_index,
                float(n.get("lat_anchor") or lat),
                float(n.get("lon_anchor") or lon),
            )
        n["lat_anchor"] = round(float(temp["lat"]), 6)
        n["lon_anchor"] = round(float(temp["lon"]), 6)
        n["lat"] = round(float(temp["lat"]), 6)
        n["lon"] = round(float(temp["lon"]), 6)


def main() -> int:
    load_baronatos_overlay()
    xlsx = pick_readable_xlsx()
    if not xlsx:
        raise SystemExit("missing referencia_localizacao.xlsx")

    district_index = load_district_geometries()

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws_npcs = wb["npcs"]
    ws_terr = wb["territorios"]
    ws_poi = wb["pontos_de_interesse"]

    npcs_raw = _sheet_rows(ws_npcs)
    terr_raw = _sheet_rows(ws_terr)
    poi_raw = _sheet_rows(ws_poi)
    role_idx = load_role_index_from_book_data()
    ghoul_master_map = load_ghoul_master_map_from_source()

    npcs: list[dict[str, Any]] = []
    by_spread: dict[str, list[dict[str, Any]]] = defaultdict(list)
    anchor_load: dict[str, int] = {}
    for row in npcs_raw:
        name = fix_mojibake(str(row.get("nome") or "").strip())
        if not name:
            continue
        region = fix_mojibake(str(row.get("regiao") or "").strip())
        group = fix_mojibake(str(row.get("grupo") or "").strip())
        clan = fix_mojibake(str(row.get("cla_ou_tipo") or "").strip())
        tier = fix_mojibake(str(row.get("tier") or "").strip())
        cluster = fix_mojibake(str(row.get("cluster_territorial") or "").strip())
        subgroup = fix_mojibake(str(row.get("subgrupo_cap8") or "").strip())
        role_info = role_idx.get(canon_name_key(name), {})
        faction = infer_faction(group)
        npc_type = infer_npc_type(group, clan)
        allowed = allowed_districts_for_npc(region, faction, cluster, clan, district_index)
        forced = forced_districts_for_npc(name, faction, clan, district_index)
        if forced:
            allowed = forced
        lat_a, lon_a, bucket = choose_anchor_for_npc(
            region,
            faction,
            cluster,
            tier,
            npc_type,
            anchor_load,
            allowed,
            district_index,
            name,
        )

        portrait_stem = str(role_info.get("file_stem") or "").strip() or stem_from_name(name)
        npc = {
            "id": f"npc_{slug(name)}",
            "name": name,
            "region": region,
            "group": group,
            "subgroup": subgroup,
            "faction": faction,
            "clan": clan if clan else "-",
            "tier": tier if tier else "-",
            "npc_type": npc_type,
            "role": str(role_info.get("role") or "").strip(),
            "domain_text": str(role_info.get("domain_text") or "").strip(),
            "cluster_territorial": cluster,
            "territory_bucket": bucket,
            "lat_anchor": lat_a,
            "lon_anchor": lon_a,
            "lat": lat_a,
            "lon": lon_a,
            "_allowed_districts": sorted(allowed),
            "_forced_districts": sorted(forced),
            "portrait_stem": portrait_stem,
        }
        npcs.append(npc)
        # Dispersao global por faccao/tipo evita aglomeracao sem depender de bucket legado.
        # Tremere/Hecata mantem permissao de proximidade relativa por clã.
        clan_n = norm(clan)
        if faction == "Camarilla" and npc_type == "Vampiro" and clan_n not in ("tremere", "hecata"):
            spread_key = "Camarilla|dispersao_total"
        elif faction == "Camarilla" and npc_type == "Vampiro":
            spread_key = f"{faction}|{clan_n}"
        elif faction == "Independentes" and npc_type == "Vampiro" and clan_n == "hecata":
            spread_key = "Independentes|hecata"
        else:
            spread_key = f"{faction}|{npc_type}"
        by_spread[spread_key].append(npc)

    for arr in by_spread.values():
        disperse_bucket(arr, district_index)
        for e in arr:
            allowed_set = set(e.get("_allowed_districts") or [])
            if allowed_set:
                snap_into_districts(
                    e,
                    allowed_set,
                    district_index,
                    float(e["lat_anchor"]),
                    float(e["lon_anchor"]),
                )

    # Pontos canônicos com coordenada fixa devem prevalecer.
    apply_forced_coords(npcs, district_index)

    # Ghouls com vinculo canonico de servico devem orbitar o regente vampiro.
    place_servants_near_masters(npcs, ghoul_master_map, district_index)

    for npc in npcs:
        dkey = find_district_key_for_point(float(npc["lat"]), float(npc["lon"]), district_index)
        npc["district"] = (district_index.get(dkey) or {}).get("name", "") if dkey else ""
        npc["district_key"] = dkey or ""
        npc.pop("_allowed_districts", None)
        npc.pop("_forced_districts", None)

    # POIs
    pois: list[dict[str, Any]] = []
    for row in poi_raw:
        name = str(row.get("ponto_interesse") or "").strip()
        if not name:
            continue
        region = str(row.get("regiao_referencia") or "").strip()
        fac = str(row.get("faccao_principal") or "").strip()
        lat = convert_seed(row.get("lat_seed"))
        lon = convert_seed(row.get("lon_seed"))
        if lat is None or lon is None:
            lat, lon, _ = anchor_for_region(region, infer_faction(fac), "")

        pois.append(
            {
                "id": f"poi_{slug(name)}",
                "name": name,
                "kind": str(row.get("tipo") or "").strip(),
                "region": region,
                "faction": fac,
                "actors": str(row.get("quem_atua_por_la") or "").strip(),
                "notes": str(row.get("observacoes") or "").strip(),
                "map_anchor": str(row.get("ponto_referencia_mapa") or "").strip(),
                "address": str(row.get("endereco_ancora") or "").strip(),
                "query": str(row.get("query_geocodificacao") or "").strip(),
                "map_ref_type": str(row.get("tipo_referencia_mapa") or "").strip(),
                "lat": round(float(lat), 6),
                "lon": round(float(lon), 6),
                "district": "",
                "district_key": "",
            }
        )

    for poi in pois:
        dkey = find_district_key_for_point(float(poi["lat"]), float(poi["lon"]), district_index)
        poi["district"] = (district_index.get(dkey) or {}).get("name", "") if dkey else ""
        poi["district_key"] = dkey or ""

    territories_rows: list[dict[str, str]] = []
    for row in terr_raw:
        t = str(row.get("Column1") or "").strip()
        if not t or norm(t) == "territorio":
            continue
        territories_rows.append(
            {
                "territorio": t,
                "nivel": str(row.get("Column2") or "").strip(),
                "faccao_principal": str(row.get("Column3") or "").strip(),
            }
        )

    payload = {
        "meta": {
            "source": str(xlsx.name),
            "note": "Canon do mapa derivado da tabela da raiz (npcs, territorios, pontos_de_interesse).",
            "year": 2035,
            "counts": {
                "npcs": len(npcs),
                "pois": len(pois),
                "territorios": len(territories_rows),
            },
        },
        "territories": {
            "dominant_rules": {
                "default": "Camarilla",
                "anarch_districts": ANARCH_DISTRICTS,
                "independent_districts": INDEP_DISTRICTS,
                "lupino_risk_districts": LUPINO_DISTRICTS,
            }
        },
        "npcs": npcs,
        "points_of_interest": pois,
        "territorios_sheet": territories_rows,
    }

    OUT_06.parent.mkdir(parents=True, exist_ok=True)
    OUT_06.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_DOCS.parent.mkdir(parents=True, exist_ok=True)
    OUT_DOCS.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT_06}")
    print(f"wrote {OUT_DOCS}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
